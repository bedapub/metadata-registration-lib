import csv
import datetime
import os
import re
import zipfile
from io import BytesIO

import openpyxl
import xlrd
import xlsxwriter
import xlwt


def write_file_from_denorm_data_2(f, data, file_format):
    """
    Write a file from the denorm_data_2 format (see fata_utils.py)

    Args:
        f (file): Input file (open stream)
        data (dict): data in denorm_data_2 format
        file_format (str): format of the exported file (xlsx, tsv or csv)

    Raises:
        Exception: Unsupported file_format
    """
    if file_format == "xlsx":
        wb = xlsxwriter.Workbook(f.name)
        bold = wb.add_format({"bold": True})
        ws = wb.add_worksheet()

        col_num = 0
        for header, data_list in data.items():
            ws.write(0, col_num, str(header), bold)

            for row_num, value in enumerate(data_list):
                ws.write(row_num + 1, col_num, str(value))

            col_num += 1

        wb.close()

    elif file_format in ["tsv", "csv"]:
        delimiters = {"tsv": "\t", "csv": ","}
        writer = csv.writer(
            f,
            delimiter=delimiters[file_format],
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writerow(data.keys())
        for row in zip(*data.values()):
            writer.writerow(row)

    else:
        raise Exception(f"Format '{file_format}' not supported")


#################################################
######## Read files as records
#################################################
def get_records_and_headers_from_csv(input_file, delimiter=","):
    """
    Read excel file file to return a list of headers and records
    Columns with empty headers are ignored

    Args:
        input_file (file): Input file (.xls or .xlsx)
        delimiter (str): CSV separator

    Returns:
        tuple:
            headers (list): Headers
            records (list): List of records {header:value}
    """
    lines = [l.decode("utf-8") for l in input_file.readlines()]

    # Read headers
    headers = next(csv.reader(lines, delimiter=delimiter))

    # Remove empty headers
    headers = [h for h in headers if h]

    # Read records
    reader = csv.DictReader(lines, delimiter=delimiter)

    records = []
    for record in reader:

        for rec_header in record.keys():
            # Remove values from columns with empty headers
            if not rec_header:
                record.pop(rec_header)
        records.append(record)

    return headers, records


def get_records_and_headers_from_excel(input_file):
    """
    Read excel file file to return a list of headers and records
    Only the first sheet is read
    Columns with empty headers are ignored

    Args:
        input_file (file): Input file (.xls or .xlsx)

    Returns:
        tuple:
            headers (list): Headers
            records (list): List of records {header:value}
    """
    workbook = openpyxl.load_workbook(
        filename=input_file, read_only=True, data_only=True
    )
    sheet = workbook.worksheets[0]

    # Read headers
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(first_row)

    # Read records
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}

        for col in range(len(headers)):
            # Ignore values from empty headers
            if headers[col]:
                raw_value = row[col]

                if type(raw_value) == datetime.datetime:
                    value = raw_value.isoformat()
                elif raw_value is None:
                    value = ""
                else:
                    value = str(raw_value)

                record[headers[col]] = value

        records.append(record)

    # Remove empty headers
    headers = [h for h in headers if h]

    workbook.close()

    return headers, records


def get_records_and_headers_from_fluidigm_plt(input_file):
    """
    Read .plt file to return a list of headers and records
    .plt is an XML format generated from the Fluidigm software

    Args:
        input_file (file): Input file

    Returns:
        tuple:
            headers (list): Headers
            records (list): List of records {header:value}
    """

    raise NotImplementedError("Fluidigm .plt XML format is not implemented yet")

    # from lxml import etree

    # element.tag
    # element.attrib
    # element.nsmap
    # element.prefix
    # element.text

    ########## WHERE TO FIND WHAT?

    #### WELL
    # <a2:ContainerWell id="ref-579" xmlns:a2="http://schemas.microsoft.com/clr/nsassem/Fluidigm.BioMark.DataModel.Dispensing/DataModel%2C%20Version%3D1.1.0.0%2C%20Culture%3Dneutral%2C%20PublicKeyToken%3Dnull">
    #    <type>Any</type>
    #    <strLabel id="ref-1000">B10</strLabel> <== Well ID
    #    <nColumnIndex>10</nColumnIndex>
    #    <strRowIndex href="#ref-977"/>
    #    <nRowNumericIndex>2</nRowNumericIndex>
    #    <index>22</index>
    # </a2:ContainerWell>

    # root = etree.parse(input_file).getroot()
    # print("TAG >", root.tag)
    # print("ATTRIB >", root.attrib)
    # print("NAMESPACE >", root.nsmap)

    # for child in root:
    #     print("-- TAG >", child.tag)
    #     print("-- ATTRIB >", child.attrib)
    #     print("-- NAMESPACE >", child.nsmap)

    #     # Example: Find all Detector names
    #     for child2 in child:
    #         if child2.tag.endswith("Detector"):
    #             print("\t-- END TAG >", child2.tag.split("}")[-1])
    #             # print("\t-- ATTRIB >", child2.attrib)
    #             # print("\t-- NAMESPACE >", child2.nsmap)

    #             for child3 in child2:
    #                 print("\t\t-- TAG = ", child3.tag)
    #                 if child3.tag == "Name":
    #                     print("\t\t\tText = ", child3.text)

    # TODO: Figure out how findall work
    # print(root.findall("WellContent"))
    # print(root.findall("SOAP-ENC:arrayType", root.nsmap))
    # print(root.findall("arrayType", root.nsmap))

    # headers_map = {}

    # headers = []
    # records = []

    # return headers, records


def get_records_and_headers_from_fluidigm_csv(input_file, delimiter=","):
    """
    Read CSV file from Fluidigm to return a list of headers and records

    Args:
        input_file (file): Input file (.xls or .xlsx)
        delimiter (str): CSV separator

    Returns:
        tuple:
            headers (list): Headers
            records (list): List of records {header:value}
    """
    lines = [l.decode("utf-8") for l in input_file.readlines()]
    reader_headers = csv.reader(lines[10:12], delimiter=delimiter)

    # Read headers (on 2 lines)
    headers = []
    headers_1 = next(reader_headers)
    headers_2 = next(reader_headers)

    for header_tuple in zip(headers_1, headers_2):
        header = " ".join([h for h in header_tuple if not h in ("", None)])
        headers.append(header)

    # Remove empty headers
    headers = [h for h in headers if h]

    headers_map = {
        "Chamber ID": "Readout ID",
        "Sample Name": "Sample IDs",
        "Sample Type": "Sample type",
        "Sample rConc": "Dilution",
        "FAM-MGB Name": "qPCR Assay (Target name)",
        "FAM-MGB Type": "Experiment Type",
        "Ct Value": "Ct Value",
        "Ct Calibrated rConc": "Ct Calibrated rConc",
        "Ct Quality": "Ct Quality",
        "Ct Call": "Ct Call",
        "Ct Threshold": "Ct Threshold",
        "Defined Comments": "Readout comment",
    }

    headers_mapped = [headers_map.get(h, h) for h in headers if h in headers_map]

    # Read records
    reader = csv.DictReader(lines[12:], fieldnames=headers_mapped, delimiter=delimiter)

    records = []
    for record in reader:

        for rec_header in record.keys():
            # Remove values from columns with empty headers
            if not rec_header:
                record.pop(rec_header)
        records.append(record)

    return headers_mapped, records


#################################################
######## Read files as list of rows
#################################################
def get_rows_from_excel_file(
    input_file,
    sheet_number=None,
    sheet_name=None,
    convert_to_str=True,
    unmerge_cells=True,
):
    """
    Extract records from file as a list of dicts

    Args:
        input_file (file): Input Excel file (.xls or .xlsx)
        sheet_number (int, optional): Sheet number. Defaults to None.
        sheet_name (str, optional): Sheet name. Defaults to None.
        convert_to_str (bool, optional): Convert all values to strings if True.
            Defaults to True.
        unmerge_cells (bool, optional): Unmerge all cells if True. Defaults to True.

    Returns:
        list: Records list of dicts
    """
    try:
        workbook = openpyxl.load_workbook(
            filename=input_file, read_only=False, data_only=True
        )
        mode = "xlsx"
    except:
        input_file.seek(0, 0)
        workbook = xlrd.open_workbook(file_contents=input_file.read())
        mode = "xls"

    if sheet_number is not None:
        sheet = find_sheet_by_number(workbook, sheet_number, mode)
    elif sheet_name is not None:
        sheet = find_sheet_by_name(workbook, sheet_name, mode)
    else:
        sheet = find_sheet_by_number(workbook, 0, mode)

    # Handle merged cells
    if unmerge_cells:
        sheet = unmerge_cells_in_sheet(sheet, mode)

    # Read actual sample data
    rows = []
    for row in gen_rows_as_list_from_excel_sheet(sheet, start_row=0, mode=mode):
        if convert_to_str:
            row = [str(v) for v in row]
        rows.append(row)

    if mode == "xlsx":
        workbook.close()

    return rows


def unmerge_cells_in_sheet(sheet, mode):
    """
    Un-merge cells and copy the same value to all sub-cells

    Args:
        sheet (sheet): Excel sheet
        mode (str): "xlsx" or "xls"

    Raises:
        NotImplementedError: Not implemented for .xls format

    Returns:
        sheet object: Sheet object with cells unmerged
    """
    if mode == "xlsx":
        merged_cell_groups = [g for g in sheet.merged_cells.ranges]
        for cell_group in merged_cell_groups:
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(
                str(cell_group)
            )
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            # print(cell_group, top_left_cell_value)
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(
                min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
            ):
                for cell in row:
                    cell.value = top_left_cell_value
    if mode == "xls":
        # TODO: Implement unmerge for xls
        raise NotImplementedError(
            "unmerge_cells_in_sheet not implemented for mode = 'xls'"
        )

    return sheet


def find_sheet_by_number(workbook, number, mode="xlsx"):
    """
    Find sheet in workbook by number

    Args:
        workbook (workbook): Input Workbook object
        number (int): Desired sheet number
        mode (str, optional): "xlsx" or "xls". Defaults to "xlsx".

    Raises:
        Exception: If sheet not found

    Returns:
        sheet: Sheet object
    """
    try:
        if mode == "xlsx":
            sheet = workbook.worksheets[number]
        elif mode == "xls":
            sheet = workbook.sheet_by_index(number)
    except:
        raise Exception(f"Sheet number {number} not found")

    return sheet


def find_sheet_by_name(workbook, name, mode="xlsx", exact_match=False):
    """
    Find sheet in workbook by name

    Args:
        workbook (workbook): Input Workbook object
        name (str): Desired sheet name
        mode (str, optional): "xlsx" or "xls". Defaults to "xlsx".
        mode (bool, optional): If False, check if name is contained in sheet names.
            Defaults to False.

    Raises:
        Exception: If sheet not found

    Returns:
        sheet: Sheet object
    """
    name_an = re.sub(r"[\W_]+", "", name).lower()

    try:
        for sheet_name in get_sheet_names(workbook, mode):
            sheet_name_an = re.sub(r"[\W_]+", "", sheet_name).lower()

            if exact_match and name_an == sheet_name_an:
                return get_sheet_by_name(workbook, sheet_name, mode)
            elif not exact_match and name_an in sheet_name_an:
                return get_sheet_by_name(workbook, sheet_name, mode)

        else:
            raise Exception(f"Sheet name '{name}' not found")

    except:
        raise Exception(f"Sheet name '{name}' not found")


def get_sheet_names(workbook, mode):
    """
    Args:
        workbook (workbook): Workbook object
        mode (str): "xlsx" or "xls".

    Returns:
        list: Sheet names
    """
    if mode == "xlsx":
        return workbook.sheetnames
    elif mode == "xls":
        return workbook.sheet_names()


def get_sheet_by_name(workbook, sheet_name, mode):
    """
    Args:
        workbook (workbook): Workbook object
        sheet_name (str): Sheet name
        mode (str): "xlsx" or "xls".

    Returns:
        list: Sheet names
    """
    if mode == "xlsx":
        return workbook[sheet_name]
    elif mode == "xls":
        return workbook.sheet_by_name(sheet_name)


def gen_rows_as_list_from_excel_sheet(sheet, start_row, mode="xlsx"):
    """
    Generator that yields rows as list

    Args:
        sheet (sheet): Excel sheet object
        start_row (int): Starting row
        mode (str, optional): "xlsx" or "xls". Defaults to "xlsx".

    Yields:
        list: Row as list
    """

    if mode == "xlsx":
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            yield list(row)

    elif mode == "xls":
        for row_num in range(start_row - 1, sheet.nrows):
            yield sheet.row_values(row_num)


def remove_empty_rows(rows, empty_values=[None, "", "None"]):
    """
    Removes empty rows from input

    Args:
        rows (iterable): Input rows
        empty_values (list, optional): What is considered Empty.
            Defaults to [None, "", "None"].

    Returns:
        list: Cleaned rows
    """
    clean_rows = []
    for row in rows:
        if not all(v in empty_values for v in row):
            clean_rows.append(row)
    return clean_rows


#################################################
######## Write files
#################################################
def write_dict_list_xlsx(file, data, headers):
    """
    Write XLSX files

    Args:
        file (file): Open file stream
        data (list): Data to write as list of mappings
        headers (list): list of strings (headers)
    """
    wb = xlsxwriter.Workbook(file.name)
    ws = wb.add_worksheet()

    # Formats
    f_header = wb.add_format({"bold": True})

    # Write headers
    for col_num, header in enumerate(headers):
        ws.write(0, col_num, header, f_header)

    # Write data
    for row_num, data_dict in enumerate(data, 1):

        for col_num, header in enumerate(headers):
            ws.write(row_num, col_num, data_dict.get(header, ""))

    # Resize columns
    for num_col, header in enumerate(headers):
        width = len(header) * 0.95 if len(header) > 10 else len(header)
        ws.set_column(num_col, num_col, width + 3)

    wb.close()


def write_dict_list_xls(file, data, headers):
    """
    Write XLS (old) files

    Args:
        file (file): Open file stream
        data (list): Data to write as list of mappings
        headers (list): list of strings (headers)
    """
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet 1")

    # Formats
    f_header = xlwt.easyxf("font: bold True;")

    # Write headers
    for col_num, header in enumerate(headers):
        ws.write(0, col_num, header, f_header)

    # Write data
    for row_num, data_dict in enumerate(data, 1):

        for col_num, header in enumerate(headers):
            ws.write(row_num, col_num, data_dict.get(header, ""))

    # Resize columns
    for num_col, header in enumerate(headers):
        width = len(header) * 0.95 if len(header) > 10 else len(header)
        ws.col(num_col).width = int((width + 3) * 400)

    wb.save(file.name)


def get_memory_zip(output_dir, file_names_to_zip):
    """
    Creates a ZIP file in memory with the given files in output_dir

    Args:
        output_dir (tempfile.TemporaryDirectory): Output directory
        file_names_to_zip (iterable): Relative file names to zip inside output_dir

    Returns:
        BytesIO: ZIPed file in memory
    """
    memory_zip_file = BytesIO()
    with zipfile.ZipFile(memory_zip_file, "w") as zf:
        for file_name in file_names_to_zip:
            os.chdir(output_dir.name)
            zf.write(file_name, compress_type=zipfile.ZIP_DEFLATED)
    memory_zip_file.seek(0)
    return memory_zip_file
