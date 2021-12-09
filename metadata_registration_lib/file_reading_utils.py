import csv
import datetime
import re

import openpyxl
import xlrd

from metadata_registration_lib.other_utils import fix_float_issue


#################################################
######## Read files as records (specific formats)
#################################################
def get_records_and_headers_from_csv(input_file, delimiter=","):
    """
    Read excel file file to return a list of headers and records
    Columns with empty headers are ignored

    Args:
        input_file (file): Input file (.xls or .xlsx)
        delimiter (str): CSV separator

    Returns:
        tuple:
            headers (list): Headers
            records (list): List of records {header:value}
    """
    lines = [l.decode("utf-8") for l in input_file.readlines()]

    # Read headers
    headers = next(csv.reader(lines, delimiter=delimiter))

    # Remove empty headers
    headers = [h for h in headers if h]

    # Read records
    reader = csv.DictReader(lines, delimiter=delimiter)

    records = []
    for record in reader:

        for rec_header in record.keys():
            # Remove values from columns with empty headers
            if not rec_header:
                record.pop(rec_header)
        records.append(record)

    return headers, records


def get_records_and_headers_from_excel(input_file):
    """
    Read excel file file to return a list of headers and records
    Only the first sheet is read
    Columns with empty headers are ignored

    Args:
        input_file (file): Input file (.xls or .xlsx)

    Returns:
        tuple:
            headers (list): Headers
            records (list): List of records {header:value}
    """
    workbook = openpyxl.load_workbook(
        filename=input_file, read_only=True, data_only=True
    )
    sheet = workbook.worksheets[0]

    # Read headers
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(first_row)

    # Read records
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}

        for col in range(len(headers)):
            # Ignore values from empty headers
            if headers[col]:
                raw_value = row[col]

                if type(raw_value) == datetime.datetime:
                    value = raw_value.isoformat()

                elif type(raw_value) == float:
                    raw_value = fix_float_issue(raw_value)
                    value = str(raw_value)

                elif raw_value is None:
                    value = ""
                else:
                    value = str(raw_value)

                record[headers[col]] = value

        records.append(record)

    # Remove empty headers
    headers = [h for h in headers if h]

    workbook.close()

    return headers, records


#################################################
######## Read files as list of rows
#################################################
def get_rows_from_excel_file(
    input_file,
    sheet_number=None,
    sheet_name=None,
    convert_to_str=True,
    unmerge_cells=True,
):
    """
    Extract records from file as a list of dicts

    Args:
        input_file (file): Input Excel file (.xls or .xlsx)
        sheet_number (int, optional): Sheet number. Defaults to None.
        sheet_name (str, optional): Sheet name. Defaults to None.
        convert_to_str (bool, optional): Convert all values to strings if True.
            Defaults to True.
        unmerge_cells (bool, optional): Unmerge all cells if True. Defaults to True.

    Returns:
        list: Records list of dicts
    """
    try:
        workbook = openpyxl.load_workbook(
            filename=input_file, read_only=False, data_only=True
        )
        mode = "xlsx"
    except:
        input_file.seek(0, 0)
        workbook = xlrd.open_workbook(file_contents=input_file.read())
        mode = "xls"

    if sheet_number is not None:
        sheet = find_sheet_by_number(workbook, sheet_number, mode)
    elif sheet_name is not None:
        sheet = find_sheet_by_name(workbook, sheet_name, mode)
    else:
        sheet = find_sheet_by_number(workbook, 0, mode)

    # Handle merged cells
    if unmerge_cells:
        sheet = unmerge_cells_in_sheet(sheet, mode)

    # Read actual sample data
    rows = []
    for row in gen_rows_as_list_from_excel_sheet(sheet, start_row=0, mode=mode):
        if convert_to_str:
            row = [str(v) for v in row]
        rows.append(row)

    if mode == "xlsx":
        workbook.close()

    return rows


def unmerge_cells_in_sheet(sheet, mode):
    """
    Un-merge cells and copy the same value to all sub-cells

    Args:
        sheet (sheet): Excel sheet
        mode (str): "xlsx" or "xls"

    Raises:
        NotImplementedError: Not implemented for .xls format

    Returns:
        sheet object: Sheet object with cells unmerged
    """
    if mode == "xlsx":
        merged_cell_groups = [g for g in sheet.merged_cells.ranges]
        for cell_group in merged_cell_groups:
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(
                str(cell_group)
            )
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            # print(cell_group, top_left_cell_value)
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(
                min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
            ):
                for cell in row:
                    cell.value = top_left_cell_value
    if mode == "xls":
        # TODO: Implement unmerge for xls
        raise NotImplementedError(
            "unmerge_cells_in_sheet not implemented for mode = 'xls'"
        )

    return sheet


def find_sheet_by_number(workbook, number, mode="xlsx"):
    """
    Find sheet in workbook by number

    Args:
        workbook (workbook): Input Workbook object
        number (int): Desired sheet number
        mode (str, optional): "xlsx" or "xls". Defaults to "xlsx".

    Raises:
        Exception: If sheet not found

    Returns:
        sheet: Sheet object
    """
    try:
        if mode == "xlsx":
            sheet = workbook.worksheets[number]
        elif mode == "xls":
            sheet = workbook.sheet_by_index(number)
    except:
        raise Exception(f"Sheet number {number} not found")

    return sheet


def find_sheet_by_name(workbook, name, mode="xlsx", exact_match=False):
    """
    Find sheet in workbook by name

    Args:
        workbook (workbook): Input Workbook object
        name (str): Desired sheet name
        mode (str, optional): "xlsx" or "xls". Defaults to "xlsx".
        mode (bool, optional): If False, check if name is contained in sheet names.
            Defaults to False.

    Raises:
        Exception: If sheet not found

    Returns:
        sheet: Sheet object
    """
    name_an = re.sub(r"[\W_]+", "", name).lower()

    try:
        for sheet_name in get_sheet_names(workbook, mode):
            sheet_name_an = re.sub(r"[\W_]+", "", sheet_name).lower()

            if exact_match and name_an == sheet_name_an:
                return get_sheet_by_name(workbook, sheet_name, mode)
            elif not exact_match and name_an in sheet_name_an:
                return get_sheet_by_name(workbook, sheet_name, mode)

        else:
            raise Exception(f"Sheet name '{name}' not found")

    except:
        raise Exception(f"Sheet name '{name}' not found")


def get_sheet_names(workbook, mode):
    """
    Args:
        workbook (workbook): Workbook object
        mode (str): "xlsx" or "xls".

    Returns:
        list: Sheet names
    """
    if mode == "xlsx":
        return workbook.sheetnames
    elif mode == "xls":
        return workbook.sheet_names()


def get_sheet_by_name(workbook, sheet_name, mode):
    """
    Args:
        workbook (workbook): Workbook object
        sheet_name (str): Sheet name
        mode (str): "xlsx" or "xls".

    Returns:
        list: Sheet names
    """
    if mode == "xlsx":
        return workbook[sheet_name]
    elif mode == "xls":
        return workbook.sheet_by_name(sheet_name)


def gen_rows_as_list_from_excel_sheet(sheet, start_row, mode="xlsx"):
    """
    Generator that yields rows as list

    Args:
        sheet (sheet): Excel sheet object
        start_row (int): Starting row
        mode (str, optional): "xlsx" or "xls". Defaults to "xlsx".

    Yields:
        list: Row as list
    """

    if mode == "xlsx":
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            yield list(row)

    elif mode == "xls":
        for row_num in range(start_row - 1, sheet.nrows):
            yield sheet.row_values(row_num)


def remove_empty_rows(rows, empty_values=[None, "", "None"]):
    """
    Removes empty rows from input

    Args:
        rows (iterable): Input rows
        empty_values (list, optional): What is considered Empty.
            Defaults to [None, "", "None"].

    Returns:
        list: Cleaned rows
    """
    clean_rows = []
    for row in rows:
        if not all(v in empty_values for v in row):
            clean_rows.append(row)
    return clean_rows
